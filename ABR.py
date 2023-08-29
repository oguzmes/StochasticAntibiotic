from gurobipy import GRB, Model, quicksum  # *
import time
import random
import itertools
import openpyxl as pyxl
import pandas as pd
import numpy
import matplotlib.pyplot as plt
import os
import pickle
import multiprocessing
import math
from datetime import datetime
import networkx as nx
import abr_config as config
from copy import deepcopy
import matplotlib.patches as mpatches
import matplotlib.cm as mplcm
import matplotlib.colors as colors
import argparse
import textwrap

import warnings
warnings.simplefilter(action='ignore', category=numpy.VisibleDeprecationWarning)
    

def ConvertStringToNumList(string):
    listem = string[1:-1].split(",")
    empty = list()
    for i in range(len(listem)):
        listem[i] = listem[i].strip()
        empty.append(int(listem[i]))
    return empty

def list_to_fixed(label):
    a = ""
    for i in label:
        a += str(i)
    return a


def Str2IntList(x):  # "0000" -> [0,0,0,0]
    return [int(i) for i in x]


def read_pkl(filename):
    with open(filename, "rb") as myFile:
        pulled = pickle.load(myFile)
    del myFile
    return pulled


def write_pkl(filename, dict_name, withTime = False):
    if withTime:
        filename = filename + datetime.now().strftime("%d%m%Y_%H%M%S")
    with open(filename, "wb") as myFile:
        pickle.dump(dict_name, myFile)
    del myFile


def HashingOptimizer(k, n, s):
    return (k+1)*23 + 10*(n+1) + 1000*(s+1)


def HashingEvaluator(k, n, s):
    # return k + n + s
    return (k+1)*67*89 + 10*(n+1) + 1000*(s+1) * 97 * 53



def IsNeighborLog(iLog, jLog):
    nLog = len(iLog)
    diff = 0
    for h in range(nLog):
        diff = diff + abs(iLog[h]-jLog[h])
    if diff == 1:
        return True
    else:
        return False


def IsNeighbor(i, j, nLog):
    return IsNeighborLog(ConvertBase2(i, nLog), ConvertBase2(j, nLog))


def ConvertBase10(x, base=2):
    L = len(x)
    y = 0
    z = 1
    for i in range(1, L+1):
        y += z*x[L-i]
        z *= base
    return y


def ConvertBase2(x, ND):
    x = bin(x)
    L = len(x)-2
    y = [0 for i in range(ND-L)]
    for i in range(0, L):
        y.append(int(x[2+i]))
    return y


def NormalizeMatrix(T):
    n = len(T)
    for i in range(n):
        rowSum = 0
        for j in range(n):
            rowSum = rowSum+T[i, j]

        if rowSum == 0:
            T[i, i] = 1
        else:
            for j in range(n):
                T[i, j] = T[i, j]/rowSum
    return T

def getNodebyAttribute(G, toLook):
    for k, v in G.nodes(data=True):
        for vk in toLook.keys():
            found = True
            if v[vk] != toLook[vk]:
                found = False
                break
        if found:
            return k
    return "NotFound!"


# =============================================================================
# Given ordered solution matrix calculate probability 
# of ending at genotype q starting from p
# =============================================================================
def ComputeExactObjective(p, solMatrix, T_matrices, q):
    di, N = solMatrix.shape
    d = T_matrices[0].shape[0]
    curr = p
    for n in range(N):
        u = numpy.zeros(d)
        for s in range(d):
            for si in range(di):
                k = solMatrix[si, n]
                if k != None:
                    next_distro = curr[0, s] * T_matrices[k][s, :]
                    u = u + next_distro
        curr = u
    return round(float(curr*q.T), 3)


def findMaxLayerSize(data):
    layer_sums = {}
    
    for key, value in data.items():
        layer = key.split('_')[1]
    
        if layer in layer_sums:
            layer_sums[layer] += value
        else:
            layer_sums[layer] = value
    
    max_layer = max(layer_sums, key=layer_sums.get)
    
    layer_sums[max_layer]    
    return layer_sums[max_layer]    
    


class TimeMachine:
    def __init__(self, data):
        self.readData(config.curr_dir + os.sep + data)
        self.neighbors = {str(l1): [l2 for l2 in self.labels if IsNeighborLog(
            l1, l2)] for l1 in self.labels}
        self.pos = {str(l): ConvertBase10(l) for l in self.labels}

    
    def readData(self, path):
        # path location of an excel file
        wb = pyxl.load_workbook(path)

        # get sheetnames in workbook excluding first one which is key
        sheet_list = wb.sheetnames[1:]

        # if rr2 exists get rr2 if not get rr1
        indexes_for_extraction = list()
        for i in range(len(sheet_list)):
            indexes_for_extraction.append(i)

        extracted_antibiotics = {i: sheet_list[indexes_for_extraction[i]] for i in range(
            len(indexes_for_extraction))}
        # since the sheets are in sorted order in terms of antibiotics
        # and their replication lengths I can use that algorithm
        self.K = len(extracted_antibiotics)
        drug_growth = dict()  # in t_list all data(excluding control) in my sheet will be included
        all_growth = dict()

        # label operations
        # in this part we assume all the labels are in te same order for all antibiotics
        labels = pd.read_excel(
            path, extracted_antibiotics[0], engine="openpyxl").columns.tolist()
        labels = labels[:int(len(labels)/2)]
        self.d = len(labels)
        for i in range(self.d):
            labels[i] = [int(j) for j in labels[i]]

        OinK = {}
        # label operations
        for i in range(self.K):
            df = pd.read_excel(
                path, extracted_antibiotics[i], engine="openpyxl")
            OinK[i] = df.shape[0]
            all_growth[i] = [
                df.iloc[j, :int(df.shape[1]/2)].tolist() for j in range(OinK[i])]
            drug_growth[i] = df

        self.OinK = OinK
        self.all_growth = all_growth
        self.extracted_antibiotics = extracted_antibiotics
        self.drug_growth = drug_growth
        self.labels = labels
        self.g = int(math.log2(len(self.labels)))
        self.extracted_antibiotics_name_dose = {k:v.split(" rr")[0] for k,v in extracted_antibiotics.items()}
        self.extracted_antibiotics_name = {k:v.split(" ")[0] for k,v in extracted_antibiotics.items()}
        
        # return all_growth, extracted_antibiotics, dfs, labels


    def CreateProbMatrix(self, growth, matrixType):
        n = len(growth)
        T = numpy.matrix(numpy.zeros((n, n)))
        i = 0
        for label_i in self.labels:
            j = 0
            for label_j in self.labels:
                if IsNeighborLog(label_i, label_j) == True:
                    rate = growth[j] - growth[i]
                    if rate > 0:
                        if matrixType == "epm":
                            rate = 1
                        T[ConvertBase10(label_i), ConvertBase10(
                            label_j)] = rate
                j = j + 1
            i = i + 1
        return NormalizeMatrix(T)

    # samples drug by randomly selecting growth rates
    def SampleDrug(self, antibiotic_index, matrixType, seed=-1):
        if seed != -1:
            random.seed(seed)
        SampledDrug = self.CreateProbMatrix(list(self.all_growth[antibiotic_index][random.randint(
            0, self.OinK[antibiotic_index]-1)][i] for i in range(len(self.labels))), matrixType)
        return SampledDrug



    def CreateSampleForSelectedMatrix(self, k, S, solutionMethod, matrixType, n=0, hashing_function=HashingOptimizer):
        if solutionMethod == "matrixReduction":  # returns a single matrix
            Sampled = sum(self.SampleDrug(
                k, matrixType, hashing_function(k, n, s)) for s in range(S)) / (S)
        return Sampled

    def Solve(self, initialState, n, finalState, solutionMethod, matrixType,
                      T_matrices_optimization, T_matrices_evaluator,
                      TimeLimit, foo=None):
        p = numpy.matrix([0 for increment in range(0, self.d)])
        q = numpy.matrix([0 for increment in range(0, self.d)])
        p[0, ConvertBase10(initialState)] = 1  # starting from where?
        q[0, ConvertBase10(finalState)] = 1  # where do i want to end up
        WithBBConstraint = 0 if "noBB" in solutionMethod else 1
        lessSteps = 1 if (T_matrices_optimization[0] ==
                          numpy.identity(16)).all() else 0
        if "Strong2stage" in solutionMethod:
            data = self.Strong2Stage(p, q, T_matrices_optimization, n,
                                     solutionMethod, 1, lessSteps, TimeLimit=TimeLimit)
        elif "Multistage" in solutionMethod:
            data = self.MultiStageAndTwoStage(p, q, T_matrices_optimization,
                                              n, solutionMethod, 1, lessSteps, WithBBConstraint,
                                              TimeLimit=TimeLimit)
        elif "Weak2stage" in solutionMethod:
            data = self.MultiStageAndTwoStage(p, q, T_matrices_optimization,
                                              n, solutionMethod, 0, lessSteps, WithBBConstraint,
                                              TimeLimit=TimeLimit)
    # =============================================================================   
    # For now enumaration methods does not exist  
    # =============================================================================
        elif solutionMethod == "MultistageEnumeration":
            data = None
        elif solutionMethod == "2stageEnumeration":
            data = None
        elif solutionMethod == "DP":
            f_df, y_df = self.DynamicProgramming(T_matrices_optimization, n, finalState)
            p_pos = int(numpy.where(p == 1)[1])
            initial = ConvertBase2(p_pos, self.g)
            u_df, y_simplified = self.CalculateUvaluesFromY(T_matrices_optimization,
                                                                     y_df, initial)
            data = {"Uf_solns": u_df, "solMatrix": y_simplified,
                    "optVal": round(f_df.to_numpy()[p_pos, 0], 3),
                    "elapsedTime": 0, "bbNode": 0}
        else:
            print("SolutionMethod does not exist.")

        if data["optVal"] != "TE":
            EvaVal = ComputeExactObjective(p, data["solMatrix"].to_numpy(),
                                              T_matrices_evaluator, q)
            data["EvaVal"] = EvaVal
        else:
            data["EvaVal"] = "TE"
        return data


    def DynamicProgramming(self, matrices, N, final):
        f = numpy.zeros((len(self.labels), N+1))
        f[self.pos[str(final)], N] = 1
        y = numpy.empty((len(self.labels), N))
        K = len(matrices)
        for n in range(N-1, -1, -1):
            for si in self.labels:
                k_solns = {}
                initial = self.pos[str(si)]
                for k in range(K):
                    tk = 0
                    for sn in list(numpy.where(matrices[k][initial, :] > 0)[1]):
                        tk = tk + matrices[k][initial, sn]*f[sn, n+1]
                    k_solns[k] = tk
                k_solns = {k: v for k, v in sorted(
                    k_solns.items(), key=lambda item: item[1], reverse=True)}
                selected_k = next(iter(k_solns))
                y[initial, n] = int(selected_k)
                f[initial, n] = k_solns[selected_k]
        idx = [str(ConvertBase2(i, self.g)) for i in range(self.d)]
        f_df = pd.DataFrame(data=f, index=idx, columns=[i for i in range(N+1)])
        y_df = pd.DataFrame(data=y, index=idx, columns=[i for i in range(N)])
        return f_df, y_df

    def CalculateUvaluesFromY(self, matrices, y_df, initial):
        y = y_df.to_numpy()
        idx = y_df.index
        N = y.shape[1]
        cols = [i for i in range(N+1)]
        U = numpy.zeros((self.d, N+1))
        U[self.pos[str(initial)], 0] = 1
        for n in range(N):
            goeq = (U[:, n] > 0).nonzero()
            for s in numpy.nditer(goeq):
                si = ConvertBase2(int(s), self.g)
                for sn in self.neighbors[str(si)] + [si]:
                    U[self.pos[str(sn)], n+1] += matrices[y[self.pos[str(si)], n]][self.pos[str(si)], self.pos[str(sn)]]*U[self.pos[str(si)], n]
        u_df = pd.DataFrame(U, index=idx, columns=cols)
        y_simplified = {}
        temparr = numpy.where(U[:, :-1] > 0, y, None)
        y_simplified = pd.DataFrame(temparr, index=idx, columns=cols[:-1])
        return u_df, y_simplified

        # Lambda can either be a binary variable OR 1
    def AddStochasticVectorConstraint(self, model, u, Lambda):
        model.addConstr((u.sum(0, '*') == Lambda))

    def FormatModelSolutions(self, model, isMultiStage, solutionMethod, elapsedTime, p,
                             N, K, Uf_vars, Lambda):
        if model.status == -11:
            data = {"Uf_solns": "TE", "solMatrix": "TE", "optVal": "TE",
                    "elapsedTime": "TE", "bbNode": "TE"}
        else:
            # elapsedTime = round(time.time() - startTime, 3)
            dM = len(self.labels) if isMultiStage else 1
            bbNode = model.NodeCount
            optVal = round(model.objVal, 3)
            solMatrix = numpy.empty((dM, N), dtype=object)
            idx = [str(ConvertBase2(i, self.g)) for i in range(self.d)]
            cols = [i for i in range(N+1)]

            for n in range(N):
                for s in range(dM):
                    for k in range(K):
                        if solutionMethod == "Strong2stage":
                            LambdaVal = Lambda[n, k].x
                        else:
                            LambdaVal = Lambda[n, k, s].x
                        if isMultiStage:
                            if n == 0:
                                if s == numpy.where(numpy.array(p)[0, :] == 1)[0][0] and LambdaVal > 0.5:
                                    solMatrix[s][n] = k
                            else:
                                if LambdaVal > 0.5 and Uf_vars[n-1][0, s].x > 10**-3:
                                    solMatrix[s][n] = k
                        else:
                            if LambdaVal > 0.5:
                                solMatrix[s][n] = k
            solMatrix_df = pd.DataFrame(solMatrix,
                                        index=idx if isMultiStage else ["0"],
                                        columns=cols[:-1])
            # store u vars
            U_tol = 0.001
            Uf_solns = {}
            Uf_solnsMatrix = numpy.zeros((len(self.labels), N+1))
            Uf_solnsMatrix[:, 0] = numpy.squeeze(numpy.asarray(p)).T
            for m in range(N):
                temp = {}
                for ls in self.labels:
                    l = ConvertBase10(ls)
                    if solutionMethod == "Strong2stage":
                        Uval = Uf_vars[m, 0][0, l].x
                    else:
                        Uval = Uf_vars[m][0, l].x
                    if Uval > U_tol:
                        temp[str(ConvertBase2(l, self.g))] = round(
                            Uval, 3)
                Uf_solns[m] = temp
                Uf_solnsMatrix[:, m+1] = list(temp[str(ele)]
                                              if str(ele) in temp else 0 for ele in idx)
            Uf_solns_df = pd.DataFrame(Uf_solnsMatrix,
                                       index=idx,
                                       columns=cols)

            data = {"Uf_solns": Uf_solns_df, "solMatrix": solMatrix_df, "optVal": optVal,
                    "elapsedTime": elapsedTime, "bbNode": bbNode}

        return data

    def MultiStageAndTwoStage(self, p, q, T_matricesIn, M, solutionMethod, isMultiStage=True,
                              lessSteps=True, isBestBDConstraint=False,
                              usedMax=300, TimeLimit=3600):

        model = Model('TimeMachineSingleStage')
        model.setParam('OutputFlag', 0)
        model.setParam("TimeLimit", TimeLimit)
        # change this and see the observation
        model.setParam('MIPGap', 0.001)
        T_matrices = list(T_matricesIn.values())

        K = len(T_matrices)
        # initializations
        r, d = p.shape
        if isMultiStage:
            dM = d
        else:
            dM = 1
        T_matrices = list(T_matricesIn.values())

        # binary variables
        Lambda = model.addVars(M, K, dM, vtype=GRB.BINARY, name="Lambda")
        if lessSteps == True:
            # first step cannot be identity matrix
            for s in range(dM):
                Lambda[0, 0, s].ub = 0

            # optional symmetry breaking constraints
            model.addConstrs(Lambda[m, 0, s] <= Lambda[m+1, 0, s]
                             for m in range(M-1) for s in range(dM))

        if usedMax < K:
            f = model.addVars(K, vtype=GRB.BINARY)
            model.addConstr(quicksum(f[k] for k in range(K)) <= usedMax)
            model.addConstrs(quicksum(Lambda[n, k, s] for s in range(
                dM) for n in range(M)) <= (d+M+100)*f[k] for k in range(K))

        # one antibiotic at a time
        model.addConstrs(quicksum(Lambda[n, k, j] for k in range(
            K)) == 1 for n in range(M) for j in range(dM))

        # product variable vectors
        Uf_vars = {(m): model.addVars(r, d, vtype=GRB.CONTINUOUS, name=[
            f"u{str(m)}_{str(i)+str(j)}" for i in range(r) for j in range(d)]) for m in range(M)}
        # first step
        model.addConstrs(Uf_vars[0][0, j] == quicksum(Lambda[0, k, l if isMultiStage else 0]
                         * p[0, l]*T_matrices[k][l, j] for l in range(d) for k in range(K)) for j in range(d))

        Vf_vars = {}
        for n in range(0, M-1):
            for k in range(0, K):
                Vf_vars[n, k] = model.addVars(
                    r, d, vtype=GRB.CONTINUOUS, lb=0, name="v"+str(n)+","+str(k))

        # linearize Vf_vars[n,k][i,j]=Uf_vars[n][i,k]*Lambda[n+1,k]
        model.addConstrs(Vf_vars[n, k][i, j] <= Lambda[n+1, k, j if isMultiStage else 0]
                         for i in range(r) for n in range(M-1) for j in range(d) for k in range(K))
        model.addConstrs(Uf_vars[n][i, j] - (1-Lambda[n+1, k, j if isMultiStage else 0]) <= Vf_vars[n, k][i, j]
                         for i in range(r) for n in range(0, M-1) for j in range(d) for k in range(K))
        model.addConstrs(Vf_vars[n, k][0, j] <= Uf_vars[n][0, j]
                         for n in range(M-1) for k in range(K) for j in range(d))

        # recursive form
        model.addConstrs(Uf_vars[n+1][i, j] == quicksum(Vf_vars[n, k][0, s]*T_matrices[k][s, j]
                         for s in range(d) for k in range(K)) for j in range(d) for n in range(M-1) for i in range(r))

        # objective function
        model.setObjective(quicksum(q[0, j]*Uf_vars[M-1][0, j]
                           for j in range(d)), GRB.MAXIMIZE)

        # #BestBD siniri
        if isBestBDConstraint == True and isMultiStage:
            model.addConstrs(quicksum(Uf_vars[n][0, s]
                                      for s in range(d)) == 1 for n in range(M))

        startTime = time.time()
        model.optimize()
        elapsedTime = round(time.time() - startTime, 3)
        data = self.FormatModelSolutions(model, isMultiStage, solutionMethod,
                                         elapsedTime, p, M, K, Uf_vars, Lambda)
        del model
        return data


    def Strong2Stage(self, p, q, T_matricesIn, N, solutionMethod, S,
                     lessSteps, TimeLimit=300):
        # enviroment=Env()
        model = Model('TimeMachineMILP')
        model.setParam('OutputFlag', 0)
        # change this and see the observation
        model.setParam('MIPGapAbs', 0.001)
        model.setParam('TimeLimit', TimeLimit)

        # initializations
        r, d = p.shape
        #T_matrices = list(T_matricesIn)
        T_matrices = list(T_matricesIn.values())
        K = len(T_matrices)

        # binary variables
        Lambda = model.addVars(N, K, vtype=GRB.BINARY, name="Lambda")

        # product variable vectors
        Uf_vars = {(n, s): model.addVars(r, d, vtype=GRB.CONTINUOUS)
                   for n in range(N) for s in range(S)}

        # one matrix for each position
        model.addConstrs((Lambda.sum(n, '*') == 1)
                         for n in range(N))

        # binary variables: selection for the first position
        model.addConstrs((quicksum(Lambda[0, k]*p[i, l]*T_matrices[k][l, j] for k in range(K) for l in range(d)) == Uf_vars[0, s][i, j])
                         for i in range(r) for j in range(d) for s in range(S))

        # auxiliary variable vectors
        Vf_vars = {}
        for n in range(1, N):
            for k in range(0, K):
                for s in range(0, S):
                    Vf_vars[n-1, k, s] = model.addVars(
                        r, d, vtype=GRB.CONTINUOUS, name="v"+str(n-1)+","+str(k)+","+str(s))
                    self.AddStochasticVectorConstraint(
                        model, Vf_vars[n-1, k, s], Lambda[n, k])

        # extended formulation
        model.addConstrs((quicksum(Vf_vars[n-1, k, s][i, l]*T_matrices[k][l, j]
                                   for k in range(K) for l in range(d)) == Uf_vars[n, s][i, j])
                         for i in range(r) for j in range(d)
                         for n in range(1, N) for s in range(S))

        model.addConstrs((quicksum(Vf_vars[n-1, k, s][i, j] for k in range(K)) == Uf_vars[n-1, s][i, j])
                         for i in range(r) for j in range(d)
                         for n in range(1, N) for s in range(S))

        # objective function
        model.setObjective((1/S)*quicksum(q[0, j]*Uf_vars[N-1, s][0, j]
                           for j in range(d) for s in range(S)), GRB.MAXIMIZE)

        if lessSteps == True:
            # first step cannot be identity matrix
            Lambda[0, 0].ub = 0

            # optional symmetry breaking constraints
            model.addConstrs(Lambda[n, 0] <= Lambda[n+1, 0]
                             for n in range(N-1))

        startTime = time.time()
        model.optimize()
        elapsedTime = round(time.time() - startTime, 3)

        data = self.FormatModelSolutions(model, 0, solutionMethod, elapsedTime, p, N, K,
                                         Uf_vars, Lambda)

        del model
    #    model.__del__()
    #    file.flush()

        return data
        # return optValList, timeElapsed, bbNode, T_orderedList #for multiple solution
        
    # =============================================================================    
    # Generates antibotics by sampling and taking average
    # =============================================================================
    def GenerateMatrix(self, matrixType, useCase = "optimization", matrixSamplingSize = 2000):
        matrixFileName = f"Matrix_useCase={useCase}_type={matrixType}_s={matrixSamplingSize}"
        HashingFunction = HashingOptimizer if useCase == "optimization" else HashingEvaluator
        if os.path.isfile(os.path.join(config.matrix_files, matrixFileName)):
            print("Returning",matrixFileName,"from existing file.")
            T_matrices = read_pkl(os.path.join(config.matrix_files, matrixFileName))
        else:
            print(matrixFileName,"does not exist, generating from scratch.")

            with multiprocessing.Pool(processes=int(os.cpu_count()/2)) as pool:
                T_matrices_temp = pool.starmap(self.CreateSampleForSelectedMatrix,
                                                            itertools.product(list(self.extracted_antibiotics.keys()),
                                                                              [matrixSamplingSize], ["matrixReduction"], 
                                                                              [matrixType], [0],
                                                                              [HashingFunction]))
    
            T_matrices_temp = {k: T_matrices_temp[k] for k in range(
                len(T_matrices_temp))}

            T_matrices_temp = {**{0: numpy.matrix(numpy.eye(16))}, **{
                k+1: T_matrices_temp[k] for k in T_matrices_temp.keys()}}
            T_matrices = deepcopy(T_matrices_temp)
            write_pkl(os.path.join(config.matrix_files, matrixFileName),T_matrices)
            del T_matrices_temp

        return T_matrices
    
    
    # =============================================================================
    # Plot solution of fitness landscape
    # =============================================================================
    def PlotSolution(self, T_matrices, solution,N, initialState, final):
        
        Uf_solns = solution["Uf_solns"]
        T_solns = solution["solMatrix"].to_numpy()
        labels = self.labels
        isMultiStage = 1 if T_solns.shape[0] != 1 else 0
        
        labels_str = [list_to_fixed(label) for label in labels]
        extracted_antibiotics = self.extracted_antibiotics_name
        extracted_antibiotics = {k+1:v for k,v in extracted_antibiotics.items()}
        extracted_antibiotics[0] = "I"
        d = self.d
    
        p = numpy.matrix([0 for i in range(0, d)])
        q = numpy.matrix([0 for i in range(0, d)])
        p[0, ConvertBase10(initialState)] = 1
        q[0, ConvertBase10(final)] = 1
    
        
        ##2nd setup
        y_increment=-4
        x_increment=5
    
        labels_xpos={n-1:n*x_increment for n in range(1,N+1)}
        labels_ypos={labels_str[label]:(label*y_increment) for label in range(len(labels_str))}

    
        initialState_str = list_to_fixed(initialState)
        final_str = list_to_fixed(final)
        
        # K_list = 
        if isMultiStage == 0:
            K_first = extracted_antibiotics[T_solns[0,0]]
        else:
            K_first = extracted_antibiotics[T_solns[ConvertBase10(Str2IntList(initialState_str)), 0]]
        
            
        cm = plt.get_cmap('tab20')

        T_soln_matrix = numpy.matrix(T_solns)
        # replace nonetypes with 0 in T_soln_matrix
        T_soln_matrix[T_soln_matrix == None] = 0        
        # flatten T_solns_matrix and get unique antibiotics as integers
        uniqueAntibiotics = list(set(T_soln_matrix.flatten().tolist()[0]))

        # assign distinguishing colors to uniqueAntibiotics using colors.Normalize and ScalarMappable
        cNorm  = colors.Normalize(vmin=0, vmax= len(uniqueAntibiotics))
        scalarMap = mplcm.ScalarMappable(norm=cNorm, cmap=cm)
        colorMap={extracted_antibiotics[int(uniqueAntibiotics[i])] : scalarMap.to_rgba(i) for i in range(len(uniqueAntibiotics))}
        colorMap["I"] = "white"


        nodes_tobeadded = [(initialState_str+"_0", {"label": initialState_str, "n": 0,
                                                    "pos": (0,labels_ypos[initialState_str]),
                                                    "p":1,
                                                    "K":K_first,
                                                    "color":colorMap[K_first]})]
        node_appear_tol = 0.01
        # determine nodes to be added filter on 
        for n in range(1,N+1):
            for label, p in Uf_solns[n].iteritems():
                if p > node_appear_tol:
                    label = ConvertStringToNumList(label)
                    label_str = list_to_fixed(label)

                    if isMultiStage == 0:
                        K = extracted_antibiotics[T_solns[0,n]] if n < N else "" 
                    else:
                        K = extracted_antibiotics[T_solns[ConvertBase10(label), n]] if n < N else ""
                    if  n < N:
                        selectedColor = colorMap[K]
                    else:
                        selectedColor = "white"
                    nodes_tobeadded.append((label_str+f"_{n}", {"label": label_str, "n": n,
                                                              "pos": (labels_xpos[n-1],labels_ypos[label_str]),
                                                              "p": round(p,3),
                                                              "K": K,
                                                              "color":selectedColor}))
        # create plot and position nodes
        G = nx.DiGraph()
        G.add_nodes_from(nodes_tobeadded)
        fixed_positions = nx.get_node_attributes(G, 'pos')

        pos = nx.spring_layout(G,pos=fixed_positions,fixed = G.nodes,k=3)
        node_labels={k:(f'{G.nodes[k]["label"]}\n{G.nodes[k]["K"]}' if G.nodes[k]["n"]!=N else 
                        f'{G.nodes[k]["label"]}\n{G.nodes[k]["p"]}') for k in G.nodes.keys()}
        
        fig=plt.figure(figsize=((N+1)*x_increment,20))
        nx.draw_networkx_labels(G,pos,node_labels)
        nx.draw_networkx_nodes(G, pos)
        
        # create edges
        edges=[]
        edge_appear_tol=0.01
        for n in range(N):
            for lsi in labels:
                si=ConvertBase10(lsi)
                solIndex = si if isMultiStage else 0
                if T_solns[solIndex,n]!= None:
                    usedAntibiotic=T_solns[solIndex,n]
                    for lsj in labels:
                        sj=ConvertBase10(lsj)
                        transitionProb=round(T_matrices[usedAntibiotic][si,sj],3)
                        if transitionProb>edge_appear_tol:
                            edges.append((getNodebyAttribute(G, {"n":n,"label":list_to_fixed(ConvertBase2(si,4))}),
                                          getNodebyAttribute(G, {"n":n+1,"label":list_to_fixed(ConvertBase2(sj,4))}),transitionProb))

        # delete unnecessary edges
        edges = [edge for edge in edges if edge[0] != "NotFound!" and edge[1] != "NotFound!"]
        G.add_weighted_edges_from(edges)
        labels_on_edge=nx.get_edge_attributes(G,'weight')
        weights=tuple(labels_on_edge.values())
        # nx.draw_networkx_edge_labels(G,pos,edge_labels=labels_on_edge,label_pos=0.65)
        nx.draw_networkx(G,pos,node_size=[max(1250,v *5000) for v in nx.get_node_attributes(G,"p").values()],
                          edge_cmap=plt.cm.get_cmap('YlGn'),width=tuple(8*i for i in list(weights)),
                          with_labels=False, node_color = list(nx.get_node_attributes(G, "color").values()))
        
        plt.gca().set_facecolor("gray")
        plt.title(initialState_str+"->"+final_str+f" N={N}\nSolutionMethod={solutionMethod} MatrixType={matrixType}",
                  fontsize = 40)
        
        
        handles = [mpatches.Patch(color = v,label = k) for k,v in colorMap.items()]
        plt.legend(loc='lower center',handles=handles,markerscale = 2.0,
                   fontsize=15,bbox_transform=fig.transFigure, ncol = 20,
                   fancybox=True, shadow = True)
        
        
        outputName = f"N{N}_{initialState_str}-{final_str}_{solutionMethod}_{matrixType}_plot.png"
        
        plt.savefig(config.solution_root + os.sep +  outputName)
        print(f"Saved {outputName} under {config.solution_root}.")

        
    def SaveSolution(self, solution):
        stats = pd.DataFrame(index = ["OptVal", "EvaVal","SolutionMethod",
                                      "bbNode","elapsedTime","matrixSamplingSize"])
        stats["Info"] = [solution["optVal"], solution["EvaVal"], solutionMethod,
                         solution["bbNode"], solution["elapsedTime"], matrixSamplingSize]
        si = list_to_fixed(initialState)
        sf = list_to_fixed(finalState)
        outputName = f"N{n}_{si}-{sf}_{solutionMethod}_{matrixType}_solution.xlsx"
        fileLoc = config.solution_root + os.sep + outputName
        extracted_antibiotics = self.extracted_antibiotics_name_dose
        extracted_antibiotics = {k+1:v for k,v in extracted_antibiotics.items()}
        extracted_antibiotics[0] = "I"
        
        with pd.ExcelWriter(fileLoc, mode = "w") as writer:
            stats.to_excel(writer, sheet_name = "Info")
            solution["Uf_solns"].to_excel(writer, sheet_name = "U_values")
            solution["solMatrix"].replace(extracted_antibiotics).to_excel(writer, 
                                                                          sheet_name = "Solution")
            
        print(f"Saved {outputName} under {config.solution_root}.")
        
def parse_args():
    # fmt: off
    parser = argparse.ArgumentParser(prog = "StochasticAntibiotic",
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter,
                                     description = textwrap.dedent('''Tool used to evaluate multiplication of matrices
                                     daha fazlasi, yazilabilir'''),
                                     epilog = "Developed by O. Mesum, Assoc. Prof. B. Kocuk")
                                     
    parser.add_argument("--dataset", type = str, default = "msx255_SuppData2017_GRME_ABR.xlsx")
    parser.add_argument("--n", default= 3, nargs='?',type = int, help ="step size")
    parser.add_argument("--initialState", default="1011", nargs='?', help ="initial state selection")
    parser.add_argument("--targetState",  default="0000", nargs='?', help ="target state selection")
    parser.add_argument("--plotSolution",
                        action="store_true",default = True,  help ="use if you want to plot solution")
    parser.add_argument("--solutionMethod", default = "DP",
                        choices= ["DP", "Multistage", "Strong2stage", "Weak2stage"],
                        help = "solution method selection")
    parser.add_argument("--matrixSamplingSize", default = 10000, type=int,
                        help = "matrix sampling size selection")
    parser.add_argument("--matrixType", default = "cpm",
                        choices=["epm", "cpm"], type = str,
                         help ="matrix type selection")
    parser.add_argument("--timeLimit", default = 3600, type = int,
                        help = "time limit (seconds) for solvers")
    args = parser.parse_args()
    return args


if __name__ == '__main__':
    args = parse_args()
    TMInstance = TimeMachine(data = args.dataset) 
    n = args.n # 1, 2, 3 ...
    TimeLimit = args.timeLimit # Solver TimeLimit
    initialState = Str2IntList(args.initialState)
    finalState = Str2IntList(args.targetState)
    solutionMethod = args.solutionMethod # DP, Multistage, Strong2stage, Weak2stage    
    matrixType = args.matrixType # epm, cpm
    matrixSamplingSize = args.matrixSamplingSize
    
    T_matrices_optimization = TMInstance.GenerateMatrix(matrixType,useCase = "optimization",
                                                       matrixSamplingSize = matrixSamplingSize)
    
    T_matrices_evaluator = TMInstance.GenerateMatrix(matrixType,useCase = "evaluator",
                                                       matrixSamplingSize = matrixSamplingSize)
    
    solution = TMInstance.Solve(initialState, n, finalState, solutionMethod,
                                        matrixType, T_matrices_optimization,
                                        T_matrices_evaluator, TimeLimit)
    # save solution
    TMInstance.SaveSolution(solution)
    
    # plot and save png
    if args.plotSolution:
        TMInstance.PlotSolution(T_matrices_optimization, solution, n, initialState, finalState)
    
    
    
